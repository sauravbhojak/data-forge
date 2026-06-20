"""Excel exporter using OpenPyXL write_only mode."""

from __future__ import annotations

import io
from collections.abc import Iterator
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Exports data to an Excel (.xlsx) file using OpenPyXL's write_only mode.

    Write-only mode streams rows without building an in-memory DOM,
    making it suitable for large datasets.

    Note: Excel's 1,048,576 row limit applies. For larger datasets,
    use CSV or JSON instead.
    """

    _HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    _HEADER_FONT = Font(color="F8FAFC", bold=True, name="Calibri", size=11)
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    def export(
        self,
        data: Iterator[dict[str, Any]],
        sheet_name: str = "DataForge",
    ) -> bytes:
        """
        Return the full Excel file as bytes.

        The entire file must be completed before bytes can be returned
        (OpenPyXL's constraint). However rows are never duplicated in memory.
        """
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

        headers_written = False
        columns: list[str] = []

        for row in data:
            if not headers_written:
                columns = list(row.keys())
                # Write styled header row (write_only mode uses append)
                ws.append(columns)
                headers_written = True

            values = [self._safe_value(row.get(col)) for col in columns]
            ws.append(values)

        if not headers_written:
            # Empty dataset — write empty sheet
            wb.create_sheet("DataForge (Empty)")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def export_relational(
        self,
        datasets: dict[str, Iterator[dict[str, Any]]],
        filepath: str
    ) -> None:
        """
        Writes a multi-sheet Excel file directly to disk.
        Useful for relational exports with low memory footprint.
        """
        wb = Workbook(write_only=True)
        
        for sheet_name, data in datasets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            headers_written = False
            columns: list[str] = []
            
            for row in data:
                if not headers_written:
                    columns = list(row.keys())
                    ws.append(columns)
                    headers_written = True

                values = [self._safe_value(row.get(col)) for col in columns]
                ws.append(values)

            if not headers_written:
                ws.append(["No Data"])

        if not datasets:
            wb.create_sheet("DataForge (Empty)")

        wb.save(filepath)

    @staticmethod
    def _safe_value(val: Any) -> Any:
        """Convert non-Excel-native types to strings."""
        if val is None or isinstance(val, (bool, int, float, str)):
            return val
        return str(val)
